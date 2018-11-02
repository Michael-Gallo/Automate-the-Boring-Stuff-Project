#! python3
#multiplicationtablemaker - make multiplication table up to the value the user enters
import openpyxl,os,sys,logging
#set directory to where file is run from
os.chdir(sys.path[0])
logging.basicConfig(level = logging.DEBUG, format = ' %(levelname)s- %(message)s')
logging.disable(logging.DEBUG)
#Set commandline argument to take max value for table
max = int(sys.argv[1])
#Create a spreadsheet
wb = openpyxl.Workbook()
ws = wb.active
#Fill First Row and first column (skipping A1) to progress to max value
for i in range(1,max+1):
    ws['A'+str(i+1)]=i
    ws['A'+str(i+1)].font=openpyxl.styles.Font(bold=True)
    currentCell=ws.cell(row=1,column=i+1)
    currentCell.value=i
    currentCell.font=openpyxl.styles.Font(bold=True)
    logging.debug(i)   
#Mutliply data in between i row and b column
for i in range (2,max+2):
    for b in range(2,max+2):
        currentCol= openpyxl.utils.get_column_letter(b)+str(1)
        currentRow= 'A'+str(i)
        ws.cell(row=i,column=b).value='='+currentCol+'*'+currentRow
#Save Worksheet
wb.save('MultiplicationTable.xlsx')